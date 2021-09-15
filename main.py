import sys
from PyQt5 import QtWidgets
from PyQt5.QtWidgets import QMainWindow, QApplication, QFileDialog
from PyQt5.uic import loadUi
import os
import openpyxl


class MainWindow(QMainWindow):
    path = ""

    def __init__(self):
        super(MainWindow, self).__init__()
        loadUi("interface.ui", self)
        self.browseButtonStart.clicked.connect(self.browseButtonStartClicked)
        self.runButton.clicked.connect(self.runButtonClicked)
        self.editStart.setText("C:/")
        self.editEnd.setText("result.xlsx")

    def browseButtonStartClicked(self):
        temp = QFileDialog.getOpenFileName(self, 'Open file', filter='Файлы Excel (*.xlsx)')
        self.editStart.setText(temp[0])
        self.path = temp[0]

    def runButtonClicked(self):
        os.chdir(self.path[:self.path.rfind("/")])
        book = openpyxl.open(self.path[self.path.rfind("/") + 1:])
        book_result = openpyxl.Workbook()
        sheet_result = book_result.active
        sheet = book.active

        string = []
        # Добавление шапки
        for i in range(0, sheet.max_column):
            string.append(sheet[1][i].value)

        sheet_result.append(string)
        # Добавление шапки

        for i in range(2, sheet.max_row + 1):
            # Обнуляем строку
            string = []

            for j in range(0, sheet.max_column):
                string.append(sheet[i][j].value)

            numbers = string[7].split("\n")

            for k in numbers:
                row = string
                row[7] = k
                sheet_result.append(row)

        book_result.save(self.path[:self.path.rfind("/") + 1] + self.editEnd.toPlainText())
        book_result.close()
        book.close()
        print("Выполнено!")


if __name__ == "__main__":
    app = QApplication(sys.argv)
    mainWindow = MainWindow()
    widget = QtWidgets.QStackedWidget()
    widget.addWidget(mainWindow)
    widget.setFixedSize(550, 280)
    widget.setWindowTitle("Telephone separator")
    widget.show()
    sys.exit(app.exec_())
